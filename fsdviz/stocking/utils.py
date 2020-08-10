"""
=============================================================
~/fsdviz/fsdviz/stocking/utils.py
 Created: 28 Aug 2019 14:02:01

 DESCRIPTION:

   Utility functions used by the fsdviz stocking application

 A. Cottrill
=============================================================
"""

from openpyxl import load_workbook
from django.conf import settings
from django.db import transaction
from django.db.models import F

from ..common.models import (
    Lake,
    StrainRaw,
    Species,
    StateProvince,
    ManagementUnit,
    Agency,
    Grid10,
    FinClip,
    FishTag,
    PhysChemMark,
    CWT,
    CWTsequence,
)
from .models import StockingMethod, LifeStage, Condition, Hatchery

from ..common.utils import to_lake_dict, toChoices

# the fields from our upload template:

REQUIRED_FIELDS = [
    "stock_id",
    "lake",
    "state_prov",
    "year",
    "month",
    "day",
    "site",
    "st_site",
    "latitude",
    "longitude",
    "grid",
    "stat_dist",
    "ls_mgmt",
    "species",
    "strain",
    "no_stocked",
    "year_class",
    "stage",
    "agemonth",
    # "mark",
    # "mark_eff",
    "tag_no",
    "tag_ret",
    "length",
    "weight",
    "condition",
    "lot_code",
    "stock_meth",
    "agency",
    # "validation",
    "notes",
    # new Spring 2020:
    "hatchery",
    "agency_stock_id",
    "finclip",
    "clip_efficiency",
    "physchem_mark",
    "tag_type",
]

xlsFields2Fdviz = {
    "GLFSD_Stock_ID": "stock_id",
    "AGENCY": "agency",
    "LAKE": "lake",
    "STATE_PROV": "state_prov",
    "STAT_DIST": "stat_dist",
    "LS_MGMT": "ls_mgmt",
    "GRID_10MIN": "grid",
    "LOCATION_PRIMARY": "site",
    "LOCATION_SECONDARY": "st_site",
    "LATITUDE": "latitude",
    "LONGITUDE": "longitude",
    "YEAR": "year",
    "MONTH": "month",
    "DAY": "day",
    "STOCK_METHOD": "stock_meth",
    "SPECIES": "species",
    "STRAIN": "strain",
    "YEAR-CLASS": "year_class",
    "LIFE_STAGE": "stage",
    "AGE_MONTHS": "agemonth",
    "CWT_Number": "tag_no",
    "TAG_RETENTION": "tag_ret",
    "MEAN_LENGTH_MM": "length",
    "TOTAL_WEIGHT_KG": "weight",
    "STOCKING_MORTALITY": "condition",
    "LOT_CODE": "lot_code",
    "NUMBER_STOCKED": "no_stocked",
    "NOTES": "notes",
    # New Spring 2020:
    "Your_Agency_Stock_ID": "agency_stock_id",
    "HATCHERY": "hatchery",
    "CLIP": "finclip",
    "CLIP_EFFICIENCY": "clip_efficiency",
    "PHYS-CHEM_MARK": "physchem_mark",
    "TAG_TYPE": "tag_type",
}


def xls2dicts(data_file):
    """A helper function to read our excel file and return a list of
    dictionaries for each row in the spreadsheet.  The keys of the
    dictionaries are determined by the first row in the spreadsheet.

    """

    wb = load_workbook(filename=data_file.open(), data_only=True)
    ws = wb.worksheets[0]

    data = []

    # process just one more than we need - keeps us from reading too many
    # records into memory if they are going to be flagged anyway.
    # add to max count - one for the header row, and one to trip the
    # too many rows flag
    maxrows = settings.MAX_UPLOAD_EVENT_COUNT + 2
    key_row = settings.UPLOAD_KEY_FIELD_ROW - 1
    first_data_row = settings.UPLOAD_FIRST_DATA_ROW - 1

    for i, row in enumerate(
        ws.iter_rows(min_row=0, max_row=(maxrows + first_data_row), values_only=True)
    ):
        if i == key_row:
            # map the values from the spread sheet to our fields
            # if it can't be found in our map - just use the value
            keys = [xlsFields2Fdviz.get(x, x) for x in row]
        elif i >= first_data_row:
            vals = [x for x in row]
            # if we have a blank row - stop
            if all(v is None for v in vals):
                break
            tmp = {k: v for k, v in zip(keys, vals)}
            data.append(tmp)

    return data


def get_xls_form_choices():
    """A helper function used by events_form for provide a list of
    dictionaries that are use to populate the choice fields in the
    stocking event form.
    """

    # stat_dist and grids will be added to dictionaries that are keyed by lake:
    stat_dist = (
        ManagementUnit.objects.select_related("lake")
        .filter(primary=True)
        .values_list("lake__abbrev", "label")
    )
    stat_dist_dict = to_lake_dict(stat_dist)

    grids = Grid10.objects.select_related("lake").values_list("lake__abbrev", "grid")
    grid_dict = to_lake_dict(grids)

    choices = {
        "lakes": [(x, x) for x in Lake.objects.values_list("abbrev", flat=True)],
        "agencies": [(x, x) for x in Agency.objects.values_list("abbrev", flat=True)],
        "state_prov": [
            (x, x) for x in StateProvince.objects.values_list("abbrev", flat=True)
        ],
        "stat_dist": stat_dist_dict,
        "species": [x for x in Species.objects.values_list("abbrev", "common_name")],
        "lifestage": [
            x for x in LifeStage.objects.values_list("abbrev", "description")
        ],
        "condition": [
            (x, x) for x in Condition.objects.values_list("condition", flat=True)
        ],
        "stocking_method": [
            x for x in StockingMethod.objects.values_list("stk_meth", "description")
        ],
        "grids": grid_dict,
    }

    return choices


def get_event_model_form_choices(event):
    """A helper function used by eventsForm to provide a dictionary of
    lists that are use to populate the choice fields in the stocking
    event model form (not the xls event form). - the single biggest
    difference is that the model form requires elements to be in the
    form (id, label), while the xls form requires them to be (abbrev,
    label) - where abbrev is the value in the spreadsheet.

    """

    lakes = [
        (x[0], "{} ({})".format(x[1], x[2]))
        for x in Lake.objects.values_list("id", "lake_name", "abbrev")
    ]

    agencies = [
        (x[0], "{} ({})".format(x[1], x[2]))
        for x in Agency.objects.values_list("id", "agency_name", "abbrev")
    ]

    hatcheries = [
        (x[0], "{} ({})".format(x[1], x[2]))
        for x in Hatchery.objects.order_by("hatchery_name").values_list(
            "id", "hatchery_name", "abbrev"
        )
    ]

    state_provs = [
        (x[0], "{} ({})".format(x[1], x[2]))
        for x in StateProvince.objects.values_list("id", "name", "abbrev")
    ]

    # filtered by the lake associated with this event:
    man_units = [
        (x[0], x[1])
        for x in ManagementUnit.objects.filter(
            lake=event.lake, primary=True
        ).values_list("id", "label")
    ]

    # filtered by the lake associated with this event:
    grids = [
        (x[0], x[1])
        for x in Grid10.objects.filter(lake=event.lake).values_list("id", "grid")
    ]

    species = [
        (x[0], "{} ({})".format(x[2], x[1]))
        for x in Species.objects.values_list("id", "abbrev", "common_name")
    ]

    # filterd by the species associated with this strain:
    strains = StrainRaw.objects.filter(species=event.species).values_list(
        "id", "description", "raw_strain"
    )
    strains = sorted(strains, key=lambda x: x[2])
    strain_choices = [(x[0], "{} ({})".format(x[2], x[1])) for x in strains]

    stocking_methods = [
        (x[0], "{} ({})".format(x[2], x[1]))
        for x in StockingMethod.objects.values_list("id", "stk_meth", "description")
    ]

    lifestages = [
        (x[0], "{} ({})".format(x[2], x[1]))
        for x in LifeStage.objects.values_list("id", "abbrev", "description")
    ]

    conditions = [
        (x[0], "{} ({})".format(x[2], x[1]))
        for x in Condition.objects.values_list("id", "condition", "description")
    ]

    fin_clips = [
        (x[0], "{} ({})".format(x[1].title(), x[0]))
        for x in FinClip.objects.order_by("description").values_list(
            "abbrev", "description"
        )
    ]

    fish_tags = [
        (x[0], "{} ({})".format(x[1].title(), x[2]))
        for x in FishTag.objects.order_by("description").values_list(
            "tag_code", "description", "tag_code"
        )
    ]

    physchem_marks = [
        (x[0], "{} ({})".format(x[1].title(), x[2]))
        for x in PhysChemMark.objects.order_by("description").values_list(
            "id", "description", "mark_code"
        )
    ]

    choices = {
        "lakes": lakes,
        "agencies": agencies,
        "state_provs": state_provs,
        "species": species,
        "lifestages": lifestages,
        "conditions": conditions,
        "stocking_methods": stocking_methods,
        "grids": grids,
        "hatcheries": hatcheries,
        "strains": strain_choices,
        "managementUnits": man_units,
        "fin_clips": fin_clips,
        "fish_tags": fish_tags,
        "physchem_marks": physchem_marks,
    }

    return choices


def form2params(formdata):
    """A helper function for our form to find specific stocking events. It
    takes a dictionary (form.cleaned_data), and returns a url query string
    that contains the query parameter values. If no values are selected,
    an empty string is returned.

    Arguments:

    - `formdata`: A dictionary containing the fields:selected values
      pairs returned from the 'find_events_form'.

    """

    params = []

    for key, value in formdata.items():
        if value:
            if isinstance(value, int):
                value = str(value)

            if isinstance(value, list):
                tmp = ",".join([str(x) for x in value])
            else:
                tmp = value
            params.append("{}={}".format(key, tmp))

    if len(params) >= 1:
        return "?" + "&".join(params)
    else:
        return ""


def validate_upload(events):
    """A function to check the basic attributes of the uploaded
    spreadsheet. The function checks to make sure that the uploaded
    data has between 1 and the max number of rows, that all of the
    required fields are included, but no more, and that the upload is
    limited to a single, year, agency and lake.  If any of these
    criteria fail, the events are considered invalid (valid=False),
    and a meaningful message is returned.

    """
    valid = True
    msg = None

    if len(events) > settings.MAX_UPLOAD_EVENT_COUNT:
        valid = False
        msg = (
            "Uploaded file has too many records. Please split it into"
            + "smaller packets (e.g by species)."
        )
    elif len(events) == 0:
        valid = False
        msg = "The uploaded file does not appear to contain any stocking records!"

    elif len(set([x["agency"] for x in events])) > 1:
        valid = False
        msg = (
            "The uploaded file has more than one agency. "
            + " Data submissions are limited to a single year, species, and agency. "
        )
    elif len(set([x["lake"] for x in events])) > 1:
        valid = False
        msg = (
            "The uploaded file has more than one lake. "
            + " Data submissions are limited to a single year, species, and agency. "
        )
    elif len(set([x["year"] for x in events])) > 1:
        valid = False
        msg = (
            "The uploaded file has more than one year. "
            + " Data submissions are limited to a single year, species, and agency. "
        )

    elif len(set(events[0].keys()) - set(REQUIRED_FIELDS)) >= 1:
        valid = False
        # note - this should be a non-critical error
        flds = list(set(events[0].keys()) - set(REQUIRED_FIELDS))
        flds.sort()

        if len(flds) == 1:
            msg = (
                "The uploaded file appears to have an additional field: {}. "
                + "This field was ignored."
            ).format(flds[0])
        else:
            field_list = ", ".join(flds)
            msg = (
                "The uploaded file appears to have {} additional field(s): {}. "
                + "These fields were ignored."
            ).format(len(flds), field_list)

    elif len(set(REQUIRED_FIELDS) - set(events[0].keys())) >= 1:
        valid = False
        # note - this should be a non-critical error
        flds = list(set(REQUIRED_FIELDS) - set(events[0].keys()))
        flds.sort()
        if len(flds) == 1:
            msg = (
                "The uploaded file appears to be missing the field: {}. "
                + "This field is required in a valid data upload template."
            ).format(flds[0])
        else:
            field_list = ", ".join(flds)
            msg = (
                "The uploaded file appears to be missing the fields: {}. "
                + "These fields are required in a valid data upload template."
            ).format(field_list)

    return valid, msg


def get_choices():
    """a helper function used to create a dictionary of dictionaries
    containing the choices for each field in a stocking event form."""

    lakes = [x for x in Lake.objects.values_list("id", "abbrev")]
    lake_choices = toChoices(lakes)

    agencies = [x for x in Agency.objects.values_list("id", "abbrev")]
    agency_choices = toChoices(agencies)

    stateProvinces = [
        x for x in StateProvince.objects.values_list("id", "abbrev", "name")
    ]
    stateProv_choices = toChoices(stateProvinces)

    species = [x for x in Species.objects.values_list("id", "abbrev", "common_name")]
    species_choices = toChoices(species)

    # TODO - add strains to lookup.
    strains = StrainRaw.objects.select_related("species").values_list(
        "id", "species__abbrev", "raw_strain"
    )

    lifestages = [
        x for x in LifeStage.objects.values_list("id", "abbrev", "description")
    ]
    lifestage_choices = toChoices(lifestages)

    conditions = [x for x in Condition.objects.values_list("id", "condition")]
    condition_choices = toChoices(conditions)

    stocking_methods = [
        x for x in StockingMethod.objects.values_list("id", "stk_meth", "description")
    ]

    stocking_method_choices = toChoices(stocking_methods)

    # grid_id_lookup will be a dictionary keyed by slug:

    grids = Grid10.objects.select_related("lake").values_list(
        "id", "slug", "lake__abbrev", "grid"
    )

    # grid choices must match the values coming in from the spreadsheet.
    # no slugs!
    tmp = [[x[2], x[3]] for x in grids]
    grid_choices = to_lake_dict(tmp)

    mus = (
        ManagementUnit.objects.filter(primary=True)
        .select_related("lake")
        .values_list("id", "slug", "lake__abbrev", "label")
    )
    tmp = [[x[2], x[3]] for x in mus]
    mu_choices = to_lake_dict(tmp)

    choices = {
        "grids": grid_choices,
        "stat_dist": mu_choices,
        "lakes": lake_choices,
        "agencies": agency_choices,
        "state_prov": stateProv_choices,
        "species": species_choices,
        "lifestage": lifestage_choices,
        "condition": condition_choices,
        "stocking_method": stocking_method_choices,
    }

    return choices


def get_or_create_cwt_sequence(
    cwt_number, tag_type="cwt", manufacturer="nmt", sequence=(0, 0)
):
    """given a cwt_number, maker and type retrieve, or create and
    retrieve a cwt series.  The cwt may or may not exist, and may have
    to be created first.

    This function is used by forms that that create stocking events with assoicated cwts.

    Arguments:
    - `cwt_number`:
    - `tag_type`: either mm or sequential
    - `manufacturer`: either nmt or mm
    - `sequence`: - a two element tuple indicating the lower and upper range

    """
    cwt_series = None

    with transaction.atomic():

        cwt, created = CWT.objects.get_or_create(
            defaults={"tag_count": 0},
            cwt_number=cwt_number,
            manufacturer=manufacturer,
            tag_type=tag_type,
        )
        if created:
            cwt.save()
        if tag_type == "sequential":
            cwt_series, created = CWTsequence.objects.get_or_create(
                cwt=cwt, sequence=sequence
            )
        else:
            cwt_series, created = CWTsequence.objects.get_or_create(cwt=cwt)
        if created:
            cwt_series.save()

    return cwt_series


def get_cwt_sequence_dict(event):
    """given a stocking event, extract and compile any cwt data as a list
    of dictionaries that can be used to populate the cwt_sequence_formset.

    If there are no cwts, return an empty list. If there are cwts,
    return a dictionary of each one with keys corresponding to the
    fields of the cwt_sequence_formset.

    Arguments:
    - `event`:

    """

    ret = event.cwt_series.annotate(
        cwt_number=F("cwt__cwt_number"),
        tag_type=F("cwt__tag_type"),
        manufacturer=F("cwt__manufacturer"),
    ).values("cwt_number", "tag_type", "manufacturer", "sequence")

    # then for each cwt, we need to extract seq_start and seq_end from sequence and pop of sequence
    for x in ret:
        sequence = x.pop("sequence")
        x["sequence_start"] = sequence.lower
        x["sequence_end"] = sequence.upper
        # format the cwt number so it matches what people expect to see:
        cwt = x.pop("cwt_number")
        x["cwt_number"] = "-".join([cwt[:2], cwt[2:4], cwt[4:6]])

    return ret
